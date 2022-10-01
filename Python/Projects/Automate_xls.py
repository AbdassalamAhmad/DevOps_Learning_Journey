import openpyxl as xl

# Here we can add for loop to go through all files *.xls in dir to do the same edit.
filename1 = "Python/Projects/try.xlsx"
def process_sheet(filename1):

    wb = xl.load_workbook(filename1)
    sheet = wb ['Sheet1']
    cell = sheet ['a1']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print (cell.value)
        new_cell_value = cell.value * 0.9
        new_cell_location = sheet.cell(row, 4)
        new_cell_location.value = new_cell_value
        print (new_cell_location.value)
    wb.save('Python/Projects/try2.xlsx')
process_sheet(filename1)


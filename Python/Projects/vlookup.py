import openpyxl as xl

filename1 = "Python/Projects/vlookup.xlsx"
end_lookup_cell = 10

def vlookup(filename1, end_lookup_cell):

    wb = xl.load_workbook(filename1)
    sheet = wb['Sheet1']

    lookup_col = []
    
    for row in range(2, end_lookup_cell+1):
        lookup_col.append(sheet.cell(row, 4).value)


    for index, lookup_value in enumerate(lookup_col):
        for row in range(2, sheet.max_row + 1):

            cell_value = sheet.cell(row, 1).value

            if lookup_value == cell_value:
                price_found_value = sheet.cell(row, 3).value
                new_cell_location = sheet.cell(index+2, 5)
                new_cell_location.value = price_found_value


    wb.save('Python/Projects/vlookup_new.xlsx')


vlookup(filename1, end_lookup_cell)
